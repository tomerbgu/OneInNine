import configparser
import datetime

import pandas as pd
import pulp
from openpyxl.reader.excel import load_workbook

import cij_creator
from cij_creator import resource_path




class Model():
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')


        cij_path = resource_path(f'data/{self.config.get("files", "cij_matrix")}')
        data_path = resource_path(f'data/{self.config.get("files", "data_file")}')
        dist_path = resource_path(f'{self.config.get("files", "distances")}')

        cij_creator.main()
        cij_matrix = pd.read_csv(cij_path, index_col=0)  # Assuming the row headers are in the first column

        c = {(i, j): cij_matrix.loc[j, i] for i in cij_matrix.columns for j in cij_matrix.index}

        self.lecturers = pd.read_excel(data_path, sheet_name="lecturer_data", index_col=0)
        self.lecturers['position'] = self.lecturers['position'].apply(
            lambda x: [i.strip() for i in x.split(',')])  # to allow multiple roles per person
        self.lec_data = self.lecturers.to_dict(orient='index')

        self.organizations = pd.read_excel(data_path, sheet_name="org_data", index_col=0)
        for d in ["first_date", "second_date", "third_date"]:
            self.organizations[d] = self.organizations[d].dt.date
        for h in ["first_time", "second_time", "third_time"]:
            self.organizations[h] = [time.strftime('%H:%M') for time in self.organizations[h]]
        self.org_data = self.organizations.to_dict(orient='index')
        self.dist_data = pd.read_csv(dist_path, header=None, names=["from", "to", "meters", "seconds", "val"], skiprows=1)

        # input
        self.org_num = list(self.organizations.index)
        self.total_volunteers = list(self.lecturers.index)
        self.lec_num = [i for i in self.total_volunteers if 'Lecturer' in self.lecturers.loc[i]['position']]
        self.guide_num = [i for i in self.total_volunteers if 'Guide' in self.lecturers.loc[i]['position']]
        self.days = self.get_all_days()  # 31 days in October
        self.slots = range(1, 42)  # 41 slots of 15 min in a day

        # parameters
        self.workshop_len = 2
        self.lec_len = 1
        self.current_year = datetime.datetime.now().year

        # decision variables
        self.indices_x = [(i, j, d, s) for i in self.org_num for j in self.total_volunteers for d in self.days for s in self.slots]
        self.indices_f = [(i, j, d) for i in self.org_num for j in self.total_volunteers for d in self.days]
        self.indices_z = [(j, d) for j in self.total_volunteers for d in self.days]
        self.indices_c = [(i, j) for i in self.org_num for j in self.total_volunteers]

        # converting slots to hours for a constraint
        start_hour = 8
        slot_interval = 15
        self.time_slots = {}
        self.inv_time_slots = dict()
        hour = start_hour
        minute = 0
        for slot in self.slots:
            if minute == 60:
                hour += 1
                minute = 0

            self.time_slots[slot] = f"{hour:02d}:{minute:02d}"
            self.inv_time_slots[f"{hour:02d}:{minute:02d}"] = slot
            minute += slot_interval

        # availability
        calendar_data_path = resource_path(f'data/{self.config.get("files", "calendar_file")}')
        workbook = load_workbook(calendar_data_path)
        sheet_names = workbook.sheetnames
        illegal_sheet_names = ['Slots']
        self.availability_df = pd.DataFrame(columns=['From', 'Until', 'Name', 'day'])
        for sheet_name in set(sheet_names) - set(illegal_sheet_names):

            lect_ava = pd.read_excel(calendar_data_path, sheet_name=sheet_name)
            # sheet_id = sheet_name.split('-', 1)[0]
            # sheet_name = sheet_name.split('-', 1)[1].strip()
            # lect_ava['id'] = sheet_id
            lect_ava['Name'] = sheet_name
            try:
                lect_ava['day'] = lect_ava['Date'].dt.date
                self.availability_df = pd.concat([self.availability_df, lect_ava])
            except:
                pass

        self.availability_df['From'] = [time.strftime('%H:%M') for time in self.availability_df['From']]
        self.availability_df['Until'] = [time.strftime('%H:%M') for time in self.availability_df['Until']]
        self.availability_df['From_index'] = self.availability_df['From'].apply(lambda x: int(self.inv_time_slots[x]))
        self.availability_df['Until_index'] = self.availability_df['Until'].apply(lambda x: int(self.inv_time_slots[x]))

        # objective function

        self.model = pulp.LpProblem("Scheduling Model", pulp.LpMaximize)
        self.x = pulp.LpVariable.dicts(name="x", indices=self.indices_x, cat=pulp.LpInteger, lowBound=0, upBound=1)
        self.z = pulp.LpVariable.dicts(name="z", indices=self.indices_z, cat=pulp.LpInteger, lowBound=0, upBound=1)
        self.f = pulp.LpVariable.dicts(name="f", indices=self.indices_f, cat=pulp.LpContinuous, lowBound=0, upBound=18)
        self.model += pulp.lpSum([c[(i, j)] * self.x[(i, j, d, s)] for (i, j, d, s) in self.indices_x])
        self.add_constraints()

    def add_constraints(self):

        for i in self.org_num:
            self.model += (pulp.lpSum([self.x[(i, j, d, s)] for d in self.days for s in self.slots for j in self.total_volunteers]) <= 1,
                      f"constraint_one_slot_per_org_{i}")

        # Constraint (9.2)
        for i in self.org_num:
            for d in self.days:
                for s in self.slots:
                    hour = self.time_slots[s]
                    # because the excel include hours and not slots, we transform the slots into hours values
                    if (d, hour) not in [(self.org_data[i]["first_date"], self.org_data[i]["first_time"]),
                                         (self.org_data[i]["second_date"], self.org_data[i]["second_time"]),
                                         (self.org_data[i]["third_date"], self.org_data[i]["third_time"])]:
                        for j in self.total_volunteers:
                            self.model += (
                                self.x[(i, j, d, s)] - 0 == 0,
                                "constraint_9.2_" + str(i) + "_" + str(j) + "_" + str(d) + "_" + str(s))

        # Constraint (1)- non-overlap constraint- one lecturer is scheduled to an organization
        for i in self.org_num:
            for d in self.days:
                for s in self.slots:
                    self.model += (pulp.lpSum([self.x[(i, j, d, s)] for j in self.total_volunteers]) <= 1,
                              "constraint_1_" + str(i) + "_" + str(d) + "_" + str(s))

        # Constraint (2)- non-overlap constraint- one organization is scheduled to a lecturer
        for j in self.total_volunteers:
            for d in self.days:
                for s in self.slots:
                    self.model += (pulp.lpSum([self.x[(i, j, d, s)] for i in self.org_num]) <= 1,
                              "constraint_2_" + str(j) + "_" + str(d) + "_" + str(s))

        # Constraint (3)- for a specific organization in a specific date- one lecturer will be scheduled
        for d in self.days:
            for i in self.org_num:
                if self.org_data[i]["is_workshop"] == 0:
                    # [org_data["id"]==i] and org_data[org_data["is_workshop"]==0]:
                    self.model += (pulp.lpSum([self.x[(i, j, d, s)] for j in self.lec_num for s in self.slots]) <= 1,
                              "constraint_3_" + str(d) + "_" + str(i))
                else:
                    self.model += (pulp.lpSum([self.x[(i, j, d, s)] for j in self.guide_num for s in self.slots]) <= 1,
                              "constraint_4_" + str(d) + "_" + str(i))

        # Constraint (5) - One lecturer per organization per available date
        for i in self.org_num:
            available_days = {self.org_data[i]["first_date"], self.org_data[i]["second_date"], self.org_data[i]["third_date"]}
            available_slots = set([slot for slot, time in self.time_slots.items() if
                                   time in [self.org_data[i]["first_time"], self.org_data[i]["second_time"],
                                            self.org_data[i]["third_time"]]])

            if self.org_data[i]["is_workshop"] == 0:
                self.model += (
                    pulp.lpSum([self.x[(i, j, d, s)] for j in self.lec_num for d in available_days for s in available_slots]) <= 1,
                    "constraint_5_" + str(i) + "_" + str(d)
                )
            else:
                self.model += (
                    pulp.lpSum([self.x[(i, j, d, s)] for j in self.guide_num for d in available_days for s in available_slots]) <= 1,
                    "constraint_6_" + str(i) + "_" + str(d)
                )

        # Constraint (7)
        for i in self.org_num:
            if self.org_data[i]["is_workshop"] == 0:
                for j in self.lec_num:
                    for d in self.days:
                        for s in self.slots:
                            self.model += (self.x[(i, j, d, s)] <= 1 - pulp.lpSum(
                                [self.x[(k, j, d, t)] for t in range(s + 1, min(s + 6, 42)) for k in self.org_num if k != i]),
                                      "constraint_7_" + str(i) + "_" + str(j) + "_" + str(d) + "_" + str(s))

        # Constraint (8)
        for i in self.org_num:
            if self.org_data[i]["is_workshop"] == 1:
                for j in self.guide_num:
                    for d in self.days:
                        for s in self.slots:
                            self.model += (
                                self.x[(i, j, d, s)] <= 1 - pulp.lpSum([self.x[(i, j, d, t)] for t in range(s + 1, min(s + 10, 42))]),
                                f"constraint_8_{j}_{i}_{d}_{s}")

        # Constraint (9) - Availability
        for j in self.total_volunteers:
            for d in self.days:
                for s in self.slots:
                    val = self.is_available(j, d, s)
                    for i in self.org_num:
                        self.model += (self.x[(i, j, d, s)] <= val,
                                  "constraint_9_" + str(i) + "_" + str(j) + "_" + str(d) + "_" + str(s))

        # Constraint (9.1)
        for i in self.org_num:
            for j in self.total_volunteers:
                self.model += (
                    pulp.lpSum(self.x[(i, j, d, s)] for d in self.days for s in self.slots) <= 1,
                    "constraint_9.1_" + str(i) + "_" + str(j)
                )

        # Constraint (10)
        for j in self.total_volunteers:
            for d in self.days:
                self.model += (
                    pulp.lpSum([self.x[(i, j, d, s)] for i in self.org_num for s in self.slots]) <= self.lec_data[j]["vol_limit"] + self.z[(j, d)],
                    "constraint_10_" + str(j) + "_" + str(d))
        # Create a set of valid organization combinations to iterate over
        valid_combinations = [(k, i) for k in self.org_num for i in self.org_num if i != k]

        # Precompute distance values and availability for each organization
        distances = {}
        availability = {}
        for k, i in valid_combinations:
            dist = float(self.dist_data.loc[(self.dist_data['from'] == self.org_data[k]["address"]) & (
                    self.dist_data['to'] == self.org_data[i]["address"]), 'seconds'].values[0])
            distances[(k, i)] = dist
            availability[(k, i)] = (
                {self.org_data[k]["first_date"], self.org_data[k]["second_date"], self.org_data[k]["third_date"]},
                {self.org_data[i]["first_date"], self.org_data[i]["second_date"], self.org_data[i]["third_date"]},
                set([slot for slot, time in self.time_slots.items() if
                     time in [self.org_data[k]["first_time"], self.org_data[k]["second_time"], self.org_data[k]["third_time"]]]),
                set([slot for slot, time in self.time_slots.items() if
                     time in [self.org_data[i]["first_time"], self.org_data[i]["second_time"], self.org_data[i]["third_time"]]])
            )

        # Constraint (11)
        for k, i in valid_combinations:
            for j in self.total_volunteers:
                days_i, days_k, slots_i, slots_k = availability[(k, i)]
                for d in days_i & days_k:
                    for s in slots_i & slots_k:
                        for o in range(1, s):
                            # Lecturer or guide
                            lec_length = self.lec_len if j in self.lec_num else self.workshop_len

                            dist = distances[(k, i)]

                            self.model += (
                                self.f[(i, j, d)] >= self.f[(k, j, d)] + lec_length + 0.5 * float(
                                    (1 - self.lec_data[j]["mobility"])) * dist
                                + dist - 10000 * (2 - self.x[(k, j, d, o)] - self.x[(i, j, d, s)])
                                , "constraint_11_" + str(k) + "_" + str(i) + "_" + str(j) + "_" + str(d) + "_" + str(
                                    s) + "_" + str(o)
                            )

        # Constraint (12)
        for i in self.org_num:
            for j in self.total_volunteers:
                for d in self.days:
                    self.model += (
                    self.f[(i, j, d)] == pulp.lpSum([self.x[(i, j, d, s)] * (8 + ((s - 1) * 15) / 60) for s in self.slots]),
                              "constraint_12_" + str(i) + "_" + str(j) + "_" + str(d))


    def solve_model(self):
        # running
        print("Solving Optimization Problem:")
        self.model.solve()

        # Check the status of the solution
        if self.model.status == pulp.LpStatusOptimal:
            # Retrieve and print the optimal objective value
            print("Total Cost = ", pulp.value(self.model.objective))

            # Retrieve and print the values of x[(i, j, d, s)]
            for item in self.indices_x:
                if (self.x[item].varValue != 0.0):
                    print("x-" + str(item), self.x[item].varValue)

            # Retrieve and print the values of u[(i, j, d)]
            for item in self.indices_f:
                if (self.f[item].varValue != 0.0):
                    print("f-" + str(item), self.f[item].varValue)

        else:
            print("Optimization problem did not find an optimal solution.")
            return None


        # Create a DataFrame to store the results
        results_df = pd.DataFrame(
            columns=['Organization', 'Volunteer', 'Date', 'Start Time', 'End Time', 'Location', 'Type', 'Confirmed'])

        # Iterate over the self.indices_x and retrieve the values
        for item in self.indices_x:
            if self.x[item].varValue != 0.0:
                org, volunteer, day, slot = item
                day_str = f"{day}/10/{self.current_year}"
                Start_time = self.time_slots[slot]
                results_df = pd.concat([results_df, pd.DataFrame({'Organization': [org],
                                                                  'Volunteer': [volunteer],
                                                                  'Date': [day_str],
                                                                  'Start Time': [Start_time],
                                                                  'End Time': [self.time_slots[slot + (
                                                                      8 if self.org_data[org]['is_workshop'] == 1 else 4)]],
                                                                  'Location': [self.org_data[org]['address']],
                                                                  'Type': 'Workshop' if self.org_data[org][
                                                                                            'is_workshop'] == 1 else 'Lecture',
                                                                  'Confirmed': 'No'})])

        print("Finished Calculations")
        return results_df

    def is_available(self, lecturer, day, time_slot):
        # function needs to return 0 if not available, 1 otherwise
        filtered_availability_df = self.availability_df[(self.availability_df['Name'] == lecturer) &
                                                   (self.availability_df['day'] == day) &
                                                   (self.availability_df['From_index'] <= time_slot) &
                                                   (self.availability_df['Until_index'] >= time_slot)]
        if filtered_availability_df.empty:
            return 0
        else:
            return 1

    def add_custom_no_match_constraints(self, constraints):
        for entry in constraints:
            i = entry["org"]
            j = entry["lec"]
            d = int(entry['date'].split('/')[0])
            s = self.inv_time_slots[entry['slot']]
            self.model += (self.x[(i, j, d, s)] == 0, f"custom_constraint_{i}_{j}_{d}_{s}")

    def add_custom_already_matched_constraints(self, constraints):
        for entry in constraints:
            i = entry["org"]
            j = entry["lec"]
            d = int(entry['date'].split('/')[0])
            s = self.inv_time_slots[entry['slot']]
            self.model += (self.x[(i, j, d, s)] == 1, f"custom_constraint_{i}_{j}_{d}_{s}")

    def get_all_days(self):
        date_set = set()
        for d in ["first_date", "second_date", "third_date"]:
            date_set.update(self.organizations[d])

        return date_set



if __name__ == '__main__':
    print("Starting calculation of best way to assign lecturers to organizations")
    model = Model()
    results_df = model.solve_model()
    # Save the results to an Excel file
    print("Saving results to file")
    results_file = model.config.get('files', 'output_file')
    results_df.to_excel(f'{results_file}.xlsx', index=False)